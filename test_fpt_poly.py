"""
test_fpt_poly.py
================
Lọc spam dữ liệu Social Listening FPT Polytechnic (Cao đẳng FPT) — 12 rules.

Logic cốt lõi (inverted filter):
  - Nhắc "FPT" / hệ sinh thái FPT nhưng KHÔNG có ngữ cảnh "Cao đẳng FPT" → SPAM
  - Topic Poly K-beauty: nhắc Cao đẳng FPT nhưng KHÔNG có ngữ cảnh K-beauty → SPAM

Usage:
    python test_fpt_poly.py <input_file.xlsx> [--output <output_file.xlsx>]
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. CONSTANTS
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = ["Id", "ParentId", "Type", "Title", "Content", "Description", "SiteName"]

# Ngữ cảnh "Cao đẳng FPT" — điều kiện để KHÔNG bị đánh spam bởi các rule FPT
CAODANG_FPT_KW = [
    "cao đẳng fpt", "fpt polytechnic", "fpt poly", "poly fpt",
    "caodangfpt", "cd fpt", "polytechnic fpt",
    "trường cao đẳng fpt", "cđ fpt",
]

# R5 – FPT University / Đại học FPT
FPT_UNIVERSITY_KW = [
    "đại học fpt", "fpt university", "fpt edu", "đhfpt", "dh fpt",
    "university fpt", "fpt university hanoi", "fpt university hcm",
    "đại học fpt hà nội", "đại học fpt hcm", "đại học fpt đà nẵng",
    "đại học fpt cần thơ", "fpt university cần thơ",
]

# R6 – FPT School
FPT_SCHOOL_KW = [
    "fpt school", "fpt schooling", "trường fpt school",
    "fptschool", "fpt k12",
]

# R7 – FPT Telecom
FPT_TELECOM_KW = [
    "fpt telecom", "fpt internet", "fpt fiber", "mạng fpt",
    "đường truyền fpt", "wifi fpt", "fpt broadband",
    "lắp mạng fpt", "gói cước fpt", "fpt isp",
]

# R8 – FPT Play
FPT_PLAY_KW = [
    "fpt play", "fptplay", "fpt cinema", "rạp chiếu fpt",
    "fpt play box", "fpt film", "fpt media",
]

# R4 – BH Media
BH_MEDIA_KW = [
    "bh media", "bhmedia", "bh entertainment",
]

# R2 – FPT Corporation / Tập đoàn
FPT_CORP_KW = [
    "tập đoàn fpt", "fpt corporation", "fpt group", "fpt software",
    "fpt retail", "fpt is", "fpt information system",
    "fpt japan", "fpt smart cloud", "fpt ai",
    "fpt securities", "fpt digital",
]

# R3 – Lãnh đạo / cá nhân FPT (không gắn với Cao đẳng FPT)
FPT_LEADERS_KW = [
    "hoàng việt anh", "trương gia bình", "nguyễn văn khoa",
    "bùi quang ngọc", "đỗ cao bảo",
]

# R9 – Catch-all: sub-brand FPT khác
FPT_SUBBRAND_KW = [
    "fpt software", "fpt retail", "fpt shop", "fpt camera",
    "fpt smart home", "fpt securities", "fpt digital retail",
    "fpt.vn", "fpt online",
]

# K-beauty ngữ cảnh (cho topic Poly K-beauty)
KBEAUTY_KW = [
    "k-beauty", "kbeauty", "k beauty", "poly k-beauty", "poly kbeauty",
    "làm đẹp", "thẩm mỹ", "nail", "spa",
    "trang điểm", "chăm sóc da", "skincare", "makeup",
    "mỹ phẩm", "phun xăm", "lông mày", "mi giả", "nối mi",
    "waxing", "facial", "dưỡng da", "hàn quốc làm đẹp",
    "korean beauty", "k-pop beauty", "esthetics",
    "gội đầu dưỡng sinh", "tạo mẫu tóc", "hair stylist",
    "chăm sóc sắc đẹp", "ngành làm đẹp",
]

# R12 – Nội dung giáo dục FPT Polytechnic chung (không K-beauty)
EDUCATION_KW = [
    "tuyển sinh", "xét tuyển", "học bổng", "học phí",
    "sinh viên", "học sinh", "giảng viên", "giáo viên",
    "campus", "ký túc xá", "câu lạc bộ", "clb", "sự kiện trường",
    "lễ tốt nghiệp", "bằng tốt nghiệp", "thực tập",
    "chương trình đào tạo", "ngành học", "khoa",
    "review trường", "trải nghiệm học", "học tại",
]

# ---------------------------------------------------------------------------
# 2. HELPERS
# ---------------------------------------------------------------------------

def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val)


def _has_caodang_fpt(text: str) -> bool:
    return any(kw in text for kw in CAODANG_FPT_KW)


def _has_kbeauty(text: str) -> bool:
    return any(kw in text for kw in KBEAUTY_KW)


def _has_fpt(text: str) -> bool:
    return "fpt" in text


# ---------------------------------------------------------------------------
# 3. CORE — 12 RULES
# ---------------------------------------------------------------------------

def check_spam(
    row: dict[str, Any],
    is_post: bool = True,
    is_kbeauty_topic: bool = False,
) -> list[str]:
    """
    Áp dụng rules FPT Polytechnic lên một dòng dữ liệu.

    Parameters
    ----------
    row              : dict với keys Title, Content, Description, ...
    is_post          : True → Post/Topic; False → Comment
    is_kbeauty_topic : True → index thuộc topic Poly K-beauty

    Returns
    -------
    list[str] — danh sách lý do spam (rỗng = NOT SPAM / fall-through)
    """
    title   = _safe_str(row.get("Title"))
    content = _safe_str(row.get("Content"))
    desc    = _safe_str(row.get("Description"))

    all_text = (title + " " + content + " " + desc).lower()

    # Nếu không nhắc đến FPT gì cả → không có căn cứ rule → fall-through
    if not _has_fpt(all_text) and not is_kbeauty_topic:
        return []

    has_cd_fpt  = _has_caodang_fpt(all_text)
    has_kbeauty = _has_kbeauty(all_text)
    reasons: list[str] = []

    # ------------------------------------------------------------------
    # R5: FPT University / Đại học FPT
    # ------------------------------------------------------------------
    if any(k in all_text for k in FPT_UNIVERSITY_KW) and not has_cd_fpt:
        reasons.append("R5: Nội dung về FPT University, không liên quan Cao đẳng FPT")

    # ------------------------------------------------------------------
    # R6: FPT School
    # ------------------------------------------------------------------
    if any(k in all_text for k in FPT_SCHOOL_KW) and not has_cd_fpt:
        reasons.append("R6: Nội dung về FPT School, không liên quan Cao đẳng FPT")

    # ------------------------------------------------------------------
    # R7: FPT Telecom
    # ------------------------------------------------------------------
    if any(k in all_text for k in FPT_TELECOM_KW) and not has_cd_fpt:
        reasons.append("R7: Nội dung về FPT Telecom, không liên quan Cao đẳng FPT")

    # ------------------------------------------------------------------
    # R8: FPT Play
    # ------------------------------------------------------------------
    if any(k in all_text for k in FPT_PLAY_KW) and not has_cd_fpt:
        reasons.append("R8: Nội dung về FPT Play, không liên quan Cao đẳng FPT")

    # ------------------------------------------------------------------
    # R4: BH Media
    # ------------------------------------------------------------------
    if any(k in all_text for k in BH_MEDIA_KW) and not has_cd_fpt:
        reasons.append("R4: Nội dung về BH Media, không liên quan Cao đẳng FPT")

    # ------------------------------------------------------------------
    # R2: Tập đoàn FPT / FPT Corporation
    # ------------------------------------------------------------------
    if any(k in all_text for k in FPT_CORP_KW) and not has_cd_fpt:
        reasons.append("R2: Nội dung về Tập đoàn FPT, không liên quan Cao đẳng FPT")

    # ------------------------------------------------------------------
    # R3: Lãnh đạo / cá nhân FPT
    # ------------------------------------------------------------------
    if any(k in all_text for k in FPT_LEADERS_KW) and not has_cd_fpt:
        reasons.append("R3: Nhắc lãnh đạo FPT, không liên quan Cao đẳng FPT")

    # ------------------------------------------------------------------
    # R9: Sub-brand FPT khác (catch-all trước R1)
    # ------------------------------------------------------------------
    if any(k in all_text for k in FPT_SUBBRAND_KW) and not has_cd_fpt:
        reasons.append("R9: Nhắc sub-brand FPT, không liên quan Cao đẳng FPT")

    # ------------------------------------------------------------------
    # R1: "FPT" chung chung, không có ngữ cảnh Cao đẳng FPT
    #     Chỉ fire nếu chưa có lý do nào từ R2–R9
    # ------------------------------------------------------------------
    if not reasons and _has_fpt(all_text) and not has_cd_fpt:
        reasons.append("R1: Nhắc FPT chung chung, không liên quan Cao đẳng FPT")

    # ------------------------------------------------------------------
    # Nếu đã có lý do spam từ R1–R9 → return sớm, không cần check K-beauty
    # ------------------------------------------------------------------
    if reasons:
        return reasons

    # ------------------------------------------------------------------
    # Từ đây: content CÓ ngữ cảnh Cao đẳng FPT (has_cd_fpt = True)
    # Áp dụng thêm rules K-beauty nếu là topic Poly K-beauty
    # ------------------------------------------------------------------
    if is_kbeauty_topic:

        # R10 + R11: FPT Polytechnic mention nhưng không có K-beauty context
        if has_cd_fpt and not has_kbeauty:
            reasons.append("R10-11: Nhắc FPT Polytechnic nhưng không liên quan Poly K-beauty")
            return reasons

        # R12: Nội dung giáo dục chung Polytechnic, không có K-beauty
        if any(k in all_text for k in EDUCATION_KW) and not has_kbeauty:
            reasons.append("R12: Nội dung giáo dục FPT Polytechnic, không liên quan K-beauty")
            return reasons

    return reasons


# ---------------------------------------------------------------------------
# 4. PIPELINE
# ---------------------------------------------------------------------------

def load_data(filepath: str | Path) -> pd.DataFrame:
    df = pd.read_excel(filepath, dtype=str)
    df = df.fillna("")
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"File thiếu cột: {missing}")
    return df


def apply_labels(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_is_post"] = df["Id"].astype(str) == df["ParentId"].astype(str)

    spam_reasons: list[list[str]] = []
    for _, row in df.iterrows():
        spam_reasons.append(check_spam(row.to_dict(), is_post=bool(row["_is_post"])))

    df["_spam_reasons"] = spam_reasons

    post_spam_map: dict[str, list[str]] = {
        str(row["Id"]): row["_spam_reasons"]
        for _, row in df[df["_is_post"]].iterrows()
    }

    for idx, row in df[~df["_is_post"]].iterrows():
        parent_reasons = post_spam_map.get(str(row["ParentId"]))
        if parent_reasons is None:
            df.at[idx, "_spam_reasons"] = check_spam(row.to_dict(), is_post=False)
        elif parent_reasons:
            df.at[idx, "_spam_reasons"] = [
                f"CASCADE: Post cha là spam ({parent_reasons[0]})"
            ]

    df["Spam_Label"]  = df["_spam_reasons"].apply(lambda r: "SPAM" if r else "NOT SPAM")
    df["Spam_Reason"] = df["_spam_reasons"].apply(lambda r: " | ".join(r) if r else "")
    df.drop(columns=["_is_post", "_spam_reasons"], inplace=True)
    return df


# ---------------------------------------------------------------------------
# 5. EXPORT
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SPAM_FILL   = PatternFill("solid", fgColor="FFE0E0")
_SPAM_FONT   = Font(color="CC0000", bold=True)


def _style_sheet(ws, df_sheet: pd.DataFrame, mark_spam: bool = False) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if mark_spam and "Spam_Label" in df_sheet.columns:
        spam_col_idx = list(df_sheet.columns).index("Spam_Label") + 1
        for row_idx, val in enumerate(df_sheet["Spam_Label"], start=2):
            if val == "SPAM":
                for col_idx in range(1, len(df_sheet.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = _SPAM_FILL
                    if col_idx == spam_col_idx:
                        cell.font = _SPAM_FONT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(_safe_str(c.value)) for c in col_cells), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def export_result(df: pd.DataFrame, output_path: str | Path) -> None:
    df_spam     = df[df["Spam_Label"] == "SPAM"].copy()
    df_not_spam = df[df["Spam_Label"] == "NOT SPAM"].copy()
    spam_count  = len(df_spam)
    total       = len(df)
    spam_pct    = spam_count / total * 100 if total else 0

    from collections import Counter
    all_reasons: list[str] = []
    for r_str in df_spam["Spam_Reason"]:
        all_reasons.extend(r.split(":")[0].strip() for r in r_str.split(" | ") if r)
    rule_counts = Counter(all_reasons)

    summary_rows = [
        {"Chỉ số": "Tổng dòng", "Giá trị": total},
        {"Chỉ số": "SPAM",      "Giá trị": spam_count},
        {"Chỉ số": "NOT SPAM",  "Giá trị": len(df_not_spam)},
        {"Chỉ số": "% SPAM",    "Giá trị": f"{spam_pct:.1f}%"},
        {"Chỉ số": "---",       "Giá trị": "--- Chi tiết theo Rule ---"},
        *[{"Chỉ số": rule, "Giá trị": cnt} for rule, cnt in sorted(rule_counts.items())],
    ]
    if spam_pct > 40:
        print(f"⚠️  CẢNH BÁO: Tỉ lệ SPAM = {spam_pct:.1f}%")

    df_summary = pd.DataFrame(summary_rows)
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer,          sheet_name="All Data", index=False)
        df_spam.to_excel(writer,     sheet_name="SPAM",     index=False)
        df_not_spam.to_excel(writer, sheet_name="NOT SPAM", index=False)
        df_summary.to_excel(writer,  sheet_name="Summary",  index=False)

    wb = load_workbook(str(output_path))
    _style_sheet(wb["All Data"],  df,          mark_spam=True)
    _style_sheet(wb["SPAM"],      df_spam,     mark_spam=True)
    _style_sheet(wb["NOT SPAM"],  df_not_spam, mark_spam=False)
    _style_sheet(wb["Summary"],   df_summary,  mark_spam=False)
    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# 6. PUBLIC API
# ---------------------------------------------------------------------------

def classify(
    title: str,
    content: str,
    description: str = "",
    is_post: bool = True,
    is_kbeauty_topic: bool = False,
) -> dict:
    """
    Phân loại spam cho một request đơn lẻ.

    Parameters
    ----------
    is_kbeauty_topic : True nếu index thuộc topic Poly K-beauty
    """
    row = {"Title": title, "Content": content, "Description": description}
    reasons = check_spam(row, is_post=is_post, is_kbeauty_topic=is_kbeauty_topic)
    return {
        "is_spam":    bool(reasons),
        "reasons":    reasons,
        "reason_str": " | ".join(reasons),
    }


def run(input_path: str | Path, output_path: str | Path | None = None) -> pd.DataFrame:
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.parent / f"{input_path.stem}_spam_labeled.xlsx"
    output_path = Path(output_path)

    print(f"📂 Đọc file: {input_path}")
    df = load_data(input_path)
    print(f"   → {len(df):,} dòng")

    print("⚙️  Áp dụng rules FPT Polytechnic...")
    df_labeled = apply_labels(df)

    spam_count = (df_labeled["Spam_Label"] == "SPAM").sum()
    total      = len(df_labeled)
    print(f"   → SPAM: {spam_count:,} / {total:,} ({spam_count/total*100:.1f}%)")

    print(f"💾 Xuất file: {output_path}")
    export_result(df_labeled, output_path)
    print("✅ Hoàn tất!")
    return df_labeled


# ---------------------------------------------------------------------------
# 7. CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Lọc spam FPT Polytechnic")
    p.add_argument("input")
    p.add_argument("--output", "-o", default=None)
    return p


def main(argv: list[str] | None = None) -> None:
    args = _build_parser().parse_args(argv)
    try:
        run(args.input, args.output)
    except FileNotFoundError as e:
        print(f"❌ {e}", file=sys.stderr); sys.exit(1)
    except ValueError as e:
        print(f"❌ {e}", file=sys.stderr); sys.exit(2)


if __name__ == "__main__":
    main()
