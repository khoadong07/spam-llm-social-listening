"""
test_shb.py
===========
Lọc spam dữ liệu Social Listening SHB Bank — 7 rules.

Usage:
    python test_shb.py <input_file.xlsx> [--output <output_file.xlsx>]

Dependencies:
    pip install pandas openpyxl
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. CONSTANTS
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = ["Id", "ParentId", "Type", "Title", "Content", "Description", "SiteName"]

SHB_KEYWORDS = [
    "shb", "ngân hàng shb", "saigon hanoi bank",
    "ngân hàng sài gòn hà nội", "ngân hàng sai gon ha noi",
]

# Ngữ cảnh tài chính/ngân hàng — dùng để loại trừ false positive ở R6, R8
FINANCE_CONTEXT_KW = [
    "lãi suất", "tiết kiệm", "vay vốn", "vay tiền", "tín dụng",
    "thẻ tín dụng", "thẻ atm", "tài khoản ngân hàng", "mở tài khoản",
    "chuyển khoản", "thanh toán", "ngân hàng số", "mobile banking",
    "internet banking", "ebanking", "stk ngân hàng", "số tài khoản ngân hàng",
    "shb", "ngân hàng",
]

# Finance context dành riêng cho R8c — không gồm tên ngân hàng để tránh vòng lặp
FINANCE_CONTEXT_R8C_KW = [
    "lãi suất", "tiết kiệm", "vay vốn", "vay tiền", "tín dụng",
    "thẻ tín dụng", "thẻ atm", "tài khoản ngân hàng", "mở tài khoản",
    "chuyển khoản", "ngân hàng số", "mobile banking",
    "internet banking", "ebanking",
]

# R2 – Bất động sản
BDS_KW = [
    "bất động sản", "bán nhà", "bán đất", "cho thuê nhà", "cho thuê căn hộ",
    "cho thuê mặt bằng", "căn hộ", "chung cư", "thổ cư", "nhà mặt tiền",
    "nhà phố", "đất nền", "nhà giá rẻ", "sổ pháp lý", "sổ hồng", "sổ đỏ",
    "đông dân cư", "khu vực kinh doanh", "mặt bằng kinh doanh",
    "phòng ngủ", "wc", "toilet", "nội thất đầy đủ", "hẻm", "shr",
]
BDS_RE = re.compile(
    r"ngang\s*\d|dài\s*\d|diện tích\s*\d|\d+\s*m2|\d+\s*wc|\d+\s*pn"
    r"|\d+\s*phòng ngủ|giá bán\s*\d|giá chỉ\s*\d",
    re.IGNORECASE,
)

# R3 – Donate / từ thiện có STK
DONATE_KW = [
    "lòng hảo tâm", "ủng hộ qua stk", "ủng hộ qua số tài khoản",
    "mong anh chị giúp đỡ", "a di đà phật", "nam mô",
    "bệnh hiểm nghèo", "hoàn cảnh khó khăn", "ủng hộ kênh",
    "quyên góp", "donate", "từ thiện", "thiện nguyện",
    "cầu xin", "xin giúp đỡ", "xin ủng hộ",
]
ACCOUNT_RE = re.compile(r"\d{6,}", re.IGNORECASE)

# R4 – Link rác / clickbait
SPAM_LINK_RE = re.compile(
    r"bit\.ly/|tinyurl\.com/|t\.co/|rb\.gy/|shorturl\.at/"
    r"|cutt\.ly/|ow\.ly/|goo\.gl/|tiny\.cc/",
    re.IGNORECASE,
)

# R5 – Rao vặt / thanh lý
THANHLY_KW = [
    "thanh lý", "pass lại", "pass đồ", "nhượng lại", "cần pass",
    "dọn nhà", "xả kho cá nhân", "xả kho", "cần bán gấp",
    "fix giá", "giá fix", "không kì kèo",
]
THANHLY_PRODUCT_KW = [
    # Gia dụng
    "tủ lạnh", "máy giặt", "điều hòa", "máy lạnh", "nồi cơm",
    "máy xay", "lò vi sóng", "bếp từ", "bình nóng lạnh",
    # Điện tử
    "điện thoại", "laptop", "máy tính", "iphone", "samsung", "ipad",
    "tai nghe", "airpods", "đồng hồ thông minh",
    # Thời trang
    "quần áo", "giày dép", "túi xách", "ví da", "đồng hồ",
    "áo thun", "quần jean", "váy đầm",
]

# R6 – Thể thao / bóng đá (không có ngữ cảnh tài chính)
SPORTS_KW = [
    "bóng đá", "trận đấu", "tỷ số", "ghi bàn", "penalty", "phạt đền",
    "hiệp 1", "hiệp 2", "ngoại hạng anh", "champions league",
    "world cup", "euro", "sea games", "v-league",
    "man utd", "manchester united", "real madrid", "barcelona",
    "liverpool", "chelsea", "arsenal", "psg", "juventus",
    "cầu thủ", "hlv", "trọng tài", "bàn thắng", "hat-trick",
    "ronaldo", "messi", "neymar", "mbappé",
    "bóng rổ", "tennis", "cầu lông", "bơi lội",
]
SPORTS_RE = re.compile(
    r"\b(bóng đá|trận đấu|tỷ số|ghi bàn|penalty|phạt đền"
    r"|hiệp [12]|ngoại hạng|champions league|world cup|v.league"
    r"|man utd|real madrid|barcelona|liverpool|chelsea|arsenal"
    r"|cầu thủ|\bhlv\b|trọng tài|bàn thắng|hat.trick"
    r"|ronaldo|messi|neymar)\b",
    re.IGNORECASE,
)

# R7 – Game / trò chơi điện tử
GAME_KW = [
    "livestream game", "live stream game", "review game",
    "tóm tắt game", "gameplay", "chơi game", "gaming",
    "rank game", "leo rank", "liên quân", "liên minh huyền thoại",
    "lol ", "pubg", "free fire", "valorant", "minecraft",
    "genshin", "mobile legend", "mlbb",
]
GAME_RE = re.compile(
    r"livestream\s+game|live\s*stream\s+game|review\s+game"
    r"|tóm tắt\s+game|gameplay|\bchơi game\b|\bgaming\b"
    r"|\bleo rank\b|liên quân|liên minh huyền thoại"
    r"|\bpubg\b|\bfree fire\b|\bvalorant\b|\bminecraft\b"
    r"|\bgenshin\b|mobile legend|\bmlbb\b",
    re.IGNORECASE,
)

# R8 – Không liên quan: thiệp cưới / tân gia / review truyện / bán hàng + STK
WEDDING_KW = [
    "đám cưới", "thiệp cưới", "lễ cưới", "hôn lễ", "vu quy",
    "tân hôn", "wedding", "đính hôn",
]
TANGIADM_KW = [
    "tân gia", "khai trương", "mừng nhà mới",
]
STORY_REVIEW_KW = [
    "review truyện", "tóm tắt truyện", "truyện ma", "truyện ngôn tình",
    "review phim", "tóm tắt phim", "spoiler",
]
SELLING_KW = [
    "inbox", "order", "ship", "giao hàng toàn quốc",
    "liên hệ để mua", "đặt hàng", "còn hàng",
]

# ---------------------------------------------------------------------------
# 2. HELPERS
# ---------------------------------------------------------------------------

def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val)


def _has_shb(text: str) -> bool:
    t = text.lower()
    return any(kw in t for kw in SHB_KEYWORDS)


def _has_finance_context(text: str) -> bool:
    t = text.lower()
    return any(kw in t for kw in FINANCE_CONTEXT_KW)


def _has_bds(text: str) -> bool:
    return any(k in text for k in BDS_KW) or bool(BDS_RE.search(text))


# ---------------------------------------------------------------------------
# 3. CORE — 7 RULES
# ---------------------------------------------------------------------------

def check_spam(
    row: dict[str, Any],
    is_post: bool = True,
    is_mainbrand: bool = False,
) -> list[str]:
    """
    Áp dụng 7 rules SHB lên một dòng dữ liệu.

    Parameters
    ----------
    row          : dict với keys Title, Content, Description, SiteName, ...
    is_post      : True → Post/Topic; False → Comment
    is_mainbrand : True → index là Mainbrand SHB (giữ data bóng đá)

    Returns
    -------
    list[str] — danh sách lý do spam (rỗng = NOT SPAM)
    """
    title     = _safe_str(row.get("Title"))
    content   = _safe_str(row.get("Content"))
    desc      = _safe_str(row.get("Description"))

    all_text = (title + " " + content + " " + desc).lower()

    reasons: list[str] = []

    # ------------------------------------------------------------------
    # R2: Bài đăng bất động sản + tên ngân hàng
    # ------------------------------------------------------------------
    if _has_bds(all_text) and _has_shb(all_text):
        reasons.append("R2: Bất động sản + tên ngân hàng SHB")

    # ------------------------------------------------------------------
    # R3: Kêu gọi donate / từ thiện có STK ngân hàng
    # ------------------------------------------------------------------
    has_donate = any(k in all_text for k in DONATE_KW)
    has_account = bool(ACCOUNT_RE.search(all_text)) or _has_shb(all_text)
    if has_donate and has_account:
        reasons.append("R3: Kêu gọi donate / từ thiện có STK")

    # ------------------------------------------------------------------
    # R4: Link rác / clickbait
    # ------------------------------------------------------------------
    if SPAM_LINK_RE.search(all_text):
        reasons.append("R4: Link rác / clickbait (short URL)")

    # ------------------------------------------------------------------
    # R5: Rao vặt / thanh lý đồ cá nhân
    # ------------------------------------------------------------------
    has_thanhly  = any(k in all_text for k in THANHLY_KW)
    has_product  = any(k in all_text for k in THANHLY_PRODUCT_KW)
    if has_thanhly and has_product:
        reasons.append("R5: Rao vặt / thanh lý đồ cá nhân")

    # ------------------------------------------------------------------
    # R6: Nội dung thể thao / bóng đá không liên quan
    #     Ngoại lệ: Mainbrand SHB → giữ data bóng đá
    # ------------------------------------------------------------------
    if not is_mainbrand:
        if SPORTS_RE.search(all_text) and not _has_finance_context(all_text):
            reasons.append("R6: Nội dung thể thao / bóng đá không liên quan tài chính")

    # ------------------------------------------------------------------
    # R7: Game / trò chơi điện tử không do ngân hàng tổ chức
    # ------------------------------------------------------------------
    if GAME_RE.search(all_text) and not _has_shb(all_text):
        reasons.append("R7: Nội dung game không liên quan ngân hàng")

    # ------------------------------------------------------------------
    # R8a: Thiệp cưới / tân gia
    # ------------------------------------------------------------------
    if any(k in all_text for k in WEDDING_KW + TANGIADM_KW):
        reasons.append("R8a: Thiệp cưới / tân gia")

    # ------------------------------------------------------------------
    # R8b: Review truyện / phim không liên quan
    # ------------------------------------------------------------------
    if any(k in all_text for k in STORY_REVIEW_KW):
        reasons.append("R8b: Review truyện / phim không liên quan")

    # ------------------------------------------------------------------
    # R8c: Bán hàng + STK ngân hàng không liên quan tài chính
    # ------------------------------------------------------------------
    if any(k in all_text for k in SELLING_KW) \
            and _has_shb(all_text) \
            and not any(k in all_text for k in FINANCE_CONTEXT_R8C_KW):
        reasons.append("R8c: Bán hàng + STK ngân hàng không liên quan tài chính")

    return reasons


# ---------------------------------------------------------------------------
# 4. PIPELINE: Load → Cascade → Label
# ---------------------------------------------------------------------------

def load_data(filepath: str | Path) -> pd.DataFrame:
    df = pd.read_excel(filepath, dtype=str)
    df = df.fillna("")
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"File thiếu cột: {missing}")
    return df


def apply_labels(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_is_post"] = df["Id"].astype(str) == df["ParentId"].astype(str)

    spam_reasons: list[list[str]] = []
    for _, row in df.iterrows():
        spam_reasons.append(check_spam(row.to_dict(), is_post=bool(row["_is_post"])))

    df["_spam_reasons"] = spam_reasons

    post_spam_map: dict[str, list[str]] = {
        str(row["Id"]): row["_spam_reasons"]
        for _, row in df[df["_is_post"]].iterrows()
    }

    for idx, row in df[~df["_is_post"]].iterrows():
        parent_reasons = post_spam_map.get(str(row["ParentId"]))
        if parent_reasons is None:
            df.at[idx, "_spam_reasons"] = check_spam(row.to_dict(), is_post=False)
        elif parent_reasons:
            df.at[idx, "_spam_reasons"] = [
                f"CASCADE: Post cha là spam ({parent_reasons[0]})"
            ]

    df["Spam_Label"]  = df["_spam_reasons"].apply(lambda r: "SPAM" if r else "NOT SPAM")
    df["Spam_Reason"] = df["_spam_reasons"].apply(lambda r: " | ".join(r) if r else "")
    df.drop(columns=["_is_post", "_spam_reasons"], inplace=True)
    return df


# ---------------------------------------------------------------------------
# 5. EXPORT
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SPAM_FILL   = PatternFill("solid", fgColor="FFE0E0")
_SPAM_FONT   = Font(color="CC0000", bold=True)


def _style_sheet(ws, df_sheet: pd.DataFrame, mark_spam: bool = False) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if mark_spam and "Spam_Label" in df_sheet.columns:
        spam_col_idx = list(df_sheet.columns).index("Spam_Label") + 1
        for row_idx, val in enumerate(df_sheet["Spam_Label"], start=2):
            if val == "SPAM":
                for col_idx in range(1, len(df_sheet.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = _SPAM_FILL
                    if col_idx == spam_col_idx:
                        cell.font = _SPAM_FONT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(_safe_str(c.value)) for c in col_cells), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def export_result(df: pd.DataFrame, output_path: str | Path) -> None:
    df_spam     = df[df["Spam_Label"] == "SPAM"].copy()
    df_not_spam = df[df["Spam_Label"] == "NOT SPAM"].copy()

    spam_count = len(df_spam)
    total      = len(df)
    spam_pct   = spam_count / total * 100 if total else 0

    from collections import Counter
    all_reasons: list[str] = []
    for r_str in df_spam["Spam_Reason"]:
        all_reasons.extend(r.split(":")[0].strip() for r in r_str.split(" | ") if r)

    rule_counts = Counter(all_reasons)
    summary_rows = [
        {"Chỉ số": "Tổng dòng", "Giá trị": total},
        {"Chỉ số": "SPAM",      "Giá trị": spam_count},
        {"Chỉ số": "NOT SPAM",  "Giá trị": len(df_not_spam)},
        {"Chỉ số": "% SPAM",    "Giá trị": f"{spam_pct:.1f}%"},
        {"Chỉ số": "---",       "Giá trị": "--- Chi tiết theo Rule ---"},
        *[{"Chỉ số": rule, "Giá trị": cnt} for rule, cnt in sorted(rule_counts.items())],
    ]
    df_summary = pd.DataFrame(summary_rows)

    if spam_pct > 40:
        print(f"⚠️  CẢNH BÁO: Tỉ lệ SPAM = {spam_pct:.1f}% (> 40%) — kiểm tra lại rules!")

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer,          sheet_name="All Data", index=False)
        df_spam.to_excel(writer,     sheet_name="SPAM",     index=False)
        df_not_spam.to_excel(writer, sheet_name="NOT SPAM", index=False)
        df_summary.to_excel(writer,  sheet_name="Summary",  index=False)

    wb = load_workbook(str(output_path))
    _style_sheet(wb["All Data"],  df,          mark_spam=True)
    _style_sheet(wb["SPAM"],      df_spam,     mark_spam=True)
    _style_sheet(wb["NOT SPAM"],  df_not_spam, mark_spam=False)
    _style_sheet(wb["Summary"],   df_summary,  mark_spam=False)
    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# 6. PUBLIC API
# ---------------------------------------------------------------------------

def classify(
    title: str,
    content: str,
    description: str = "",
    site_name: str = "",
    is_post: bool = True,
    is_mainbrand: bool = False,
) -> dict:
    """
    Phân loại spam cho một request đơn lẻ.

    Parameters
    ----------
    title, content, description : nội dung bài viết
    site_name    : tên site/page nguồn (hiện chưa dùng, để mở rộng)
    is_post      : True nếu là Post/Topic, False nếu là Comment
    is_mainbrand : True nếu index là Mainbrand SHB (miễn lọc bóng đá)

    Returns
    -------
    {"is_spam": bool, "reasons": list[str], "reason_str": str}
    """
    row = {
        "Title":       title,
        "Content":     content,
        "Description": description,
        "SiteName":    site_name,
    }
    reasons = check_spam(row, is_post=is_post, is_mainbrand=is_mainbrand)
    return {
        "is_spam":    bool(reasons),
        "reasons":    reasons,
        "reason_str": " | ".join(reasons),
    }


def run(input_path: str | Path, output_path: str | Path | None = None) -> pd.DataFrame:
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.parent / f"{input_path.stem}_spam_labeled.xlsx"
    output_path = Path(output_path)

    print(f"📂 Đọc file: {input_path}")
    df = load_data(input_path)
    print(f"   → {len(df):,} dòng | {df.columns.tolist()}")

    print("⚙️  Áp dụng 7 rules SHB + cascade logic...")
    df_labeled = apply_labels(df)

    spam_count = (df_labeled["Spam_Label"] == "SPAM").sum()
    total      = len(df_labeled)
    print(f"   → SPAM: {spam_count:,} / {total:,} ({spam_count/total*100:.1f}%)")

    print(f"💾 Xuất file: {output_path}")
    export_result(df_labeled, output_path)
    print("✅ Hoàn tất!")
    return df_labeled


# ---------------------------------------------------------------------------
# 7. CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Lọc spam Social Listening SHB Bank (7 rules)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("input",  help="File xlsx đầu vào")
    p.add_argument("--output", "-o", default=None)
    return p


def main(argv: list[str] | None = None) -> None:
    args = _build_parser().parse_args(argv)
    try:
        run(args.input, args.output)
    except FileNotFoundError as e:
        print(f"❌ Không tìm thấy file: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"❌ Lỗi dữ liệu: {e}", file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()
