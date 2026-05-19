"""
test_panasonic.py
=================
Lọc spam dữ liệu Social Listening Panasonic — 7 rules từ spam_rule sheet.

Usage:
    python test_panasonic.py <input_file.xlsx> [--output <output_file.xlsx>]

Dependencies:
    pip install pandas openpyxl
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. CONSTANTS
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = ["Id", "ParentId", "Type", "Title", "Content", "Description", "SiteName"]

PANASONIC_BRAND = ["panasonic", "pana"]

# R1 – Bất động sản
BDS_KW = [
    "bán nhà", "bán đất", "bán căn hộ", "bán chung cư", "bán gấp nhà",
    "cho thuê nhà", "cho thuê căn hộ", "cho thuê mặt bằng",
    "căn hộ", "chung cư", "nhà phố", "đất nền", "đất thổ cư",
    "mặt tiền", "hẻm", "ngõ", "phòng ngủ", "wc", "toilet", "phòng tắm",
    "nội thất đầy đủ", "sổ hồng", "sổ đỏ", "shr",
]
BDS_RE = re.compile(
    r"ngang\s*\d|dài\s*\d|diện tích\s*\d|\d+\s*m2|\d+\s*wc|\d+\s*pn"
    r"|\d+\s*phòng ngủ|\d+\s*phòng tắm",
    re.IGNORECASE,
)

# R2 – Thanh lý / rao bán hàng cũ
THANHLY_KW = [
    "thanh lý", "bán lại", "cần bán", "bán gấp", "cần pass", "pass lại",
    "sell", "rao bán", "cần ra đi", "xả kho",
]
USED_GOODS_KW = [
    "đã qua sử dụng", "second hand", "hàng cũ", "hàng bãi", "nội địa nhật",
    "nội địa", "zin nguyên", "zin", "mới đẹp", "còn tốt", "đã tét",
    "đã dùng", "qua sử dụng",
]
PRICE_RE = re.compile(
    r"\b\d{1,3}tr\b|\b\d{1,3}\s*triệu\b|\b\d{1,3}k\b"
    r"|\d{1,3}[.,]\d{3}[.,]\d{3}|\d{1,3}\.\d{3}đ"
    r"|giá chỉ|giá còn|giá rẻ",
    re.IGNORECASE,
)
SELL_CONTACT_KW = ["inbox", "zalo", "liên hệ", "alo", "gọi", "nhắn tin", "ib"]

# R3 – Nhật bãi / Bãi Nhật (site name hoặc content)
NHAT_BAI_RE = re.compile(
    r"nhật\s*bãi|bãi\s*nhật|hàng\s*nhật\s*bãi|đồ\s*nhật\s*bãi"
    r"|nhat\s*bai|bai\s*nhat|hang\s*nhat\s*bai",
    re.IGNORECASE,
)

# R4 – Dịch vụ sửa chữa / lắp ráp bên thứ 3
REPAIR_KW = [
    "dịch vụ sửa", "chuyên sửa", "thợ sửa", "sửa chữa", "sửa điều hòa",
    "sửa máy lạnh", "sửa tủ lạnh", "sửa máy giặt", "bảo dưỡng",
    "vệ sinh máy lạnh", "vệ sinh điều hòa", "lắp đặt điều hòa",
    "thợ lắp", "lắp máy lạnh", "lắp ráp điều hòa",
]
REPAIR_CONTACT_RE = re.compile(
    r"0[35789]\d{8}|\+84[35789]\d{8}|gọi ngay|liên hệ ngay|hotline",
    re.IGNORECASE,
)

# R5 – Số điện thoại (áp dụng cho Topic/Post)
PHONE_RE = re.compile(r"0[35789]\d{8}|\+84[35789]\d{8}|0\d{2,3}[.\-]\d{3,4}[.\-]\d{3,4}", re.IGNORECASE)

# R6 – Kênh điện máy (áp dụng cho Topic/Post)
DIEN_MAY_RE = re.compile(
    r"điện máy xanh|điện máy chợ lớn|điện máy nguyễn kim|mediamart|media mart"
    r"|\bpico\b|nguyễn kim|fpt shop|vinhpro|điện máy thiên hòa"
    r"|điện máy hoàng hải|điện máy phượng vàng|điện máy mạnh cường"
    r"|siêu thị điện máy",
    re.IGNORECASE,
)

# R7 – Minigame (áp dụng cả Post lẫn Comment)
MINIGAME_KW = [
    "minigame", "mini game", "mini-game",
    "bình luận để nhận", "comment để nhận", "bình luận để trúng",
    "tag bạn bè để nhận", "tag.*để nhận",
    "like và share", "like & share", "like share",
    "vòng quay may mắn", "vòng quay",
    "trả lời câu hỏi để nhận", "trả lời để nhận",
    "tham gia để nhận", "điền form để nhận",
]
MINIGAME_RE = re.compile(
    r"minigame|mini.?game"
    r"|bình luận để (nhận|trúng|tham gia)"
    r"|comment để (nhận|trúng)"
    r"|tag\s+\w+\s+để nhận"
    r"|like\s+(và|&|share)"
    r"|vòng quay (may mắn|quà)"
    r"|trả lời (câu hỏi )?để nhận",
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# 2. HELPERS
# ---------------------------------------------------------------------------

def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val)


def _has_pana(text: str) -> bool:
    t = text.lower()
    return any(kw in t for kw in PANASONIC_BRAND)


def _has_bds(text: str) -> bool:
    return any(k in text for k in BDS_KW) or bool(BDS_RE.search(text))


def _has_repair(text: str) -> bool:
    return any(k in text for k in REPAIR_KW)


# ---------------------------------------------------------------------------
# 3. CORE — 7 RULES
# ---------------------------------------------------------------------------

def check_spam(row: dict[str, Any], is_post: bool = True) -> list[str]:
    """
    Áp dụng 7 rules Panasonic lên một dòng dữ liệu.

    Parameters
    ----------
    row     : dict với keys Title, Content, Description, SiteName, Type, ...
    is_post : True → Post/Topic; False → Comment

    Returns
    -------
    list[str] — danh sách lý do spam (rỗng = NOT SPAM)
    """
    title     = _safe_str(row.get("Title"))
    content   = _safe_str(row.get("Content"))
    desc      = _safe_str(row.get("Description"))
    site_name = _safe_str(row.get("SiteName")).lower()

    all_text = (title + " " + content + " " + desc).lower()

    reasons: list[str] = []

    # ------------------------------------------------------------------
    # R1: Bất động sản + Panasonic
    # ------------------------------------------------------------------
    if _has_bds(all_text) and _has_pana(all_text):
        reasons.append("R1: Bất động sản + thiết bị Panasonic")

    # ------------------------------------------------------------------
    # R2: Thanh lý / rao bán sản phẩm Panasonic
    # ------------------------------------------------------------------
    is_thanhly = any(k in all_text for k in THANHLY_KW)
    is_used    = any(k in all_text for k in USED_GOODS_KW)
    has_price  = bool(PRICE_RE.search(all_text))
    has_sell_contact = any(k in all_text for k in SELL_CONTACT_KW)

    if _has_pana(all_text) and (is_thanhly or is_used) and (has_price or has_sell_contact):
        reasons.append("R2: Thanh lý / rao bán sản phẩm Panasonic")

    # ------------------------------------------------------------------
    # R3: Site name hoặc nội dung chứa "Nhật bãi" / "Bãi Nhật"
    # ------------------------------------------------------------------
    if NHAT_BAI_RE.search(site_name) or NHAT_BAI_RE.search(all_text):
        reasons.append("R3: Rao bán sản phẩm Nhật bãi / Bãi Nhật")

    # ------------------------------------------------------------------
    # R4: Dịch vụ sửa chữa / lắp ráp bên thứ 3
    # ------------------------------------------------------------------
    if _has_repair(all_text) and REPAIR_CONTACT_RE.search(all_text):
        reasons.append("R4: Dịch vụ sửa chữa / lắp ráp bên thứ 3")

    # ------------------------------------------------------------------
    # R5: Số điện thoại trong bài đăng (Post/Topic)
    #     Comment có SĐT vẫn qua → không đánh spam
    # ------------------------------------------------------------------
    if is_post and PHONE_RE.search(all_text):
        reasons.append("R5: Bài đăng (Post) có số điện thoại")

    # ------------------------------------------------------------------
    # R6: Kênh điện máy + là bài đăng (Post)
    #     Comment từ kênh điện máy vẫn alert bình thường
    # ------------------------------------------------------------------
    if is_post and DIEN_MAY_RE.search(site_name):
        reasons.append("R6: Bài đăng từ kênh điện máy")

    # ------------------------------------------------------------------
    # R7: Minigame — cả Post lẫn Comment
    # ------------------------------------------------------------------
    if MINIGAME_RE.search(all_text):
        reasons.append("R7: Nội dung liên quan Minigame")

    return reasons


# ---------------------------------------------------------------------------
# 4. PIPELINE: Load → Cascade → Label
# ---------------------------------------------------------------------------

def load_data(filepath: str | Path) -> pd.DataFrame:
    df = pd.read_excel(filepath, dtype=str)
    df = df.fillna("")
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"File thiếu cột: {missing}")
    return df


def apply_labels(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_is_post"] = df["Id"].astype(str) == df["ParentId"].astype(str)

    spam_reasons: list[list[str]] = []
    for _, row in df.iterrows():
        spam_reasons.append(check_spam(row.to_dict(), is_post=bool(row["_is_post"])))

    df["_spam_reasons"] = spam_reasons

    # Cascade: comment của post spam → cascade spam
    post_spam_map: dict[str, list[str]] = {
        str(row["Id"]): row["_spam_reasons"]
        for _, row in df[df["_is_post"]].iterrows()
    }

    for idx, row in df[~df["_is_post"]].iterrows():
        parent_reasons = post_spam_map.get(str(row["ParentId"]))
        if parent_reasons is None:
            df.at[idx, "_spam_reasons"] = check_spam(row.to_dict(), is_post=False)
        elif parent_reasons:
            df.at[idx, "_spam_reasons"] = [
                f"CASCADE: Post cha là spam ({parent_reasons[0]})"
            ]
        # else: post cha NOT SPAM → comment NOT SPAM, giữ nguyên []

    df["Spam_Label"]  = df["_spam_reasons"].apply(lambda r: "SPAM" if r else "NOT SPAM")
    df["Spam_Reason"] = df["_spam_reasons"].apply(lambda r: " | ".join(r) if r else "")
    df.drop(columns=["_is_post", "_spam_reasons"], inplace=True)
    return df


# ---------------------------------------------------------------------------
# 5. EXPORT
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SPAM_FILL   = PatternFill("solid", fgColor="FFE0E0")
_SPAM_FONT   = Font(color="CC0000", bold=True)


def _style_sheet(ws, df_sheet: pd.DataFrame, mark_spam: bool = False) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if mark_spam and "Spam_Label" in df_sheet.columns:
        spam_col_idx = list(df_sheet.columns).index("Spam_Label") + 1
        for row_idx, val in enumerate(df_sheet["Spam_Label"], start=2):
            if val == "SPAM":
                for col_idx in range(1, len(df_sheet.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = _SPAM_FILL
                    if col_idx == spam_col_idx:
                        cell.font = _SPAM_FONT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(_safe_str(c.value)) for c in col_cells), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def export_result(df: pd.DataFrame, output_path: str | Path) -> None:
    df_spam     = df[df["Spam_Label"] == "SPAM"].copy()
    df_not_spam = df[df["Spam_Label"] == "NOT SPAM"].copy()

    spam_count = len(df_spam)
    total      = len(df)
    spam_pct   = spam_count / total * 100 if total else 0

    from collections import Counter
    all_reasons: list[str] = []
    for r_str in df_spam["Spam_Reason"]:
        all_reasons.extend(r.split(":")[0].strip() for r in r_str.split(" | ") if r)

    rule_counts = Counter(all_reasons)
    summary_rows = [
        {"Chỉ số": "Tổng dòng", "Giá trị": total},
        {"Chỉ số": "SPAM",      "Giá trị": spam_count},
        {"Chỉ số": "NOT SPAM",  "Giá trị": len(df_not_spam)},
        {"Chỉ số": "% SPAM",    "Giá trị": f"{spam_pct:.1f}%"},
        {"Chỉ số": "---",       "Giá trị": "--- Chi tiết theo Rule ---"},
        *[{"Chỉ số": rule, "Giá trị": cnt} for rule, cnt in sorted(rule_counts.items())],
    ]
    df_summary = pd.DataFrame(summary_rows)

    if spam_pct > 40:
        print(f"⚠️  CẢNH BÁO: Tỉ lệ SPAM = {spam_pct:.1f}% (> 40%) — kiểm tra lại rules!")

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer,          sheet_name="All Data", index=False)
        df_spam.to_excel(writer,     sheet_name="SPAM",     index=False)
        df_not_spam.to_excel(writer, sheet_name="NOT SPAM", index=False)
        df_summary.to_excel(writer,  sheet_name="Summary",  index=False)

    wb = load_workbook(str(output_path))
    _style_sheet(wb["All Data"],  df,          mark_spam=True)
    _style_sheet(wb["SPAM"],      df_spam,     mark_spam=True)
    _style_sheet(wb["NOT SPAM"],  df_not_spam, mark_spam=False)
    _style_sheet(wb["Summary"],   df_summary,  mark_spam=False)
    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# 6. PUBLIC API — dùng trong unit test hoặc gọi trực tiếp
# ---------------------------------------------------------------------------

def classify(
    title: str,
    content: str,
    description: str = "",
    site_name: str = "",
    is_post: bool = True,
) -> dict:
    """
    Phân loại spam cho một request đơn lẻ.

    Parameters
    ----------
    title, content, description : nội dung bài viết
    site_name : tên site/page nguồn
    is_post   : True nếu là bài đăng (Topic), False nếu là Comment

    Returns
    -------
    {
        "is_spam"      : bool,
        "reasons"      : list[str],   # danh sách rule kích hoạt
        "reason_str"   : str,         # join bằng " | "
    }
    """
    row = {
        "Title":       title,
        "Content":     content,
        "Description": description,
        "SiteName":    site_name,
    }
    reasons = check_spam(row, is_post=is_post)
    return {
        "is_spam":    bool(reasons),
        "reasons":    reasons,
        "reason_str": " | ".join(reasons),
    }


def run(input_path: str | Path, output_path: str | Path | None = None) -> pd.DataFrame:
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.parent / f"{input_path.stem}_spam_labeled.xlsx"
    output_path = Path(output_path)

    print(f"📂 Đọc file: {input_path}")
    df = load_data(input_path)
    print(f"   → {len(df):,} dòng | {df.columns.tolist()}")

    print("⚙️  Áp dụng 7 rules Panasonic + cascade logic...")
    df_labeled = apply_labels(df)

    spam_count = (df_labeled["Spam_Label"] == "SPAM").sum()
    total      = len(df_labeled)
    print(f"   → SPAM: {spam_count:,} / {total:,} ({spam_count/total*100:.1f}%)")

    print(f"💾 Xuất file: {output_path}")
    export_result(df_labeled, output_path)
    print("✅ Hoàn tất!")

    return df_labeled


# ---------------------------------------------------------------------------
# 7. CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Lọc spam Social Listening Panasonic (7 rules)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("input",  help="File xlsx đầu vào")
    p.add_argument("--output", "-o", default=None,
                   help="File xlsx đầu ra (mặc định: <input>_spam_labeled.xlsx)")
    return p


def main(argv: list[str] | None = None) -> None:
    args = _build_parser().parse_args(argv)
    try:
        run(args.input, args.output)
    except FileNotFoundError as e:
        print(f"❌ Không tìm thấy file: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"❌ Lỗi dữ liệu: {e}", file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()
