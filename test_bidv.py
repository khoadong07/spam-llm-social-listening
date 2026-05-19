"""
bidv_spam_filter.py
===================
Lọc spam dữ liệu Social Listening BIDV — Version 1.4 (2026-05-18)

16 rules + cascade logic theo SKILL_loc_spam_BIDV.md

Usage:
    python bidv_spam_filter.py <input_file.xlsx> [--output <output_file.xlsx>]

Dependencies:
    pip install pandas openpyxl
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. CONSTANTS
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = ["Id", "ParentId", "Type", "Title", "Content", "Description", "Channel", "SiteName"]

BIDV_KEYWORDS: list[str] = [
    "bidv",
    "bid",
    "ngân hàng đầu tư và phát triển việt nam",
    "ngân hàng tmcp đầu tư và phát triển việt nam",
    "ngân hàng tmcp đầu tư phát triển",
    "bsc",
    "bảo hiểm bic",
    "bidvtv",
    "bidvnews",
    "smartbanking",
    "ngân hàng đt & pt việt nam",
    "bank for investment and development of vietnam",
    "trần bắc hà",
    "bank hoa mai",
    "bidv metlife",
    "bidvmetlife",
]

# ---------------------------------------------------------------------------
# 2. HELPER FUNCTIONS
# ---------------------------------------------------------------------------

def _safe_str(val: Any) -> str:
    """Chuyển giá trị bất kỳ sang str an toàn, trả về '' nếu NaN/None."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val)


def has_bidv(text: str) -> bool:
    """Trả về True nếu text chứa bất kỳ keyword BIDV nào."""
    if not text:
        return False
    t = text.lower()
    return any(kw in t for kw in BIDV_KEYWORDS)


def _desc_has_bidv_stk(desc: str) -> bool:
    """
    Kiểm tra Description có chứa BIDV kèm số tài khoản thực (6+ chữ số).
    Bỏ qua hashtag (#BIDVHome) và mã chứng khoán (#bid).
    Dùng cho R4.
    """
    if not desc:
        return False
    d_clean = re.sub(r"#\w+", "", desc.lower())
    if not any(kw in d_clean for kw in ["bidv", "ngân hàng đầu tư", "smartbanking"]):
        return False
    return bool(re.search(r"bidv.{0,30}\d{6,}|\d{6,}.{0,30}bidv", d_clean))


# ---------------------------------------------------------------------------
# 3. CORE SPAM CHECKER — 16 RULES
# ---------------------------------------------------------------------------

def check_spam(row: dict[str, Any], is_post: bool = False) -> list[str]:
    """
    Áp dụng 16 rules lên một dòng dữ liệu.

    Parameters
    ----------
    row     : dict với keys Title, Content, Description, Channel, Type, ...
    is_post : True  → Post (xét all_text = 3 cột cho hầu hết rules)
              False → Comment độc lập (giữ logic cũ, content-specific)

    Returns
    -------
    list[str] — danh sách lý do spam (rỗng = NOT SPAM)
    """
    title   = _safe_str(row.get("Title"))
    content = _safe_str(row.get("Content"))
    desc    = _safe_str(row.get("Description"))

    all_text = (title + " " + content + " " + desc).lower()
    tc_text  = (title + " " + content).lower()

    # bidv_check_text: Post → all_text, Comment → content
    bidv_check_text = all_text if is_post else content.lower()

    reasons: list[str] = []

    # ------------------------------------------------------------------
    # R1: Đấu giá + BID
    # ------------------------------------------------------------------
    if any(k in all_text for k in ["đấu giá", "phiên đấu", "#daugia", "dau gia"]) \
            and "bid" in all_text:
        reasons.append("R1: Thảo luận đấu giá + BID")

    # ------------------------------------------------------------------
    # R2: Dịch vụ địa phương / BĐS + BIDV là dấu mốc địa lý
    # ------------------------------------------------------------------
    dv_kw = [
        "cho thuê", "bán nhà", "bán đất", "mặt bằng", "căn hộ", "bất động sản",
        "thuê nhà", "thuê mặt", "phòng trọ",
        "làm tóc", "nail", "spa", "thẩm mỹ", "salon", "hair", "massage",
        "quán ăn", "nhà hàng", "quán cafe", "cafe", "dịch vụ",
        "thời trang", "#thoitrang", "áo sơ mi", "áo thun", "quần",
        "ăn trưa", "ăn tối", "ăn sáng",
        "peel", "mỹ phẩm", "chăm sóc da", "skincare", "serum", "kem dưỡng",
        "phòng tập", "gym", "fitness",
    ]
    geo_bidv = bool(re.search(
        r"(cạnh|gần|bên cạnh|đối diện|trước|sau|cách|100m|200m|50m)"
        r".{0,20}(ngân hàng\s+)?bidv"
        r"|tầng\s*\d+.{0,15}(ngân hàng\s+)?bidv"
        r"|(ngân hàng\s+)?bidv.{0,15}tầng\s*\d+",
        all_text,
    ))
    if any(k in all_text for k in dv_kw) and geo_bidv:
        reasons.append("R2: Dịch vụ/BĐS + BIDV là dấu mốc địa lý")

    # ------------------------------------------------------------------
    # R3: Tiếng nước ngoài (Trung, Nhật, Hàn, Ả Rập)
    # ------------------------------------------------------------------
    foreign_chars = re.findall(r"[一-鿿぀-ゟ゠-ヿ가-퟿؀-ۿ]", title + content)
    if len(foreign_chars) > 5:
        reasons.append("R3: Data tiếng nước ngoài")

    # ------------------------------------------------------------------
    # R4: YouTube + BIDV là STK ở Description (không phải hashtag/ticker)
    # ------------------------------------------------------------------
    is_yt = (
        _safe_str(row.get("Channel")).lower() == "youtube"
        or "youtube" in _safe_str(row.get("Type")).lower()
    )
    if is_yt and _desc_has_bidv_stk(desc) and not has_bidv(tc_text):
        reasons.append("R4: YouTube - STK BIDV ở Description, title/content không nhắc BIDV")

    # ------------------------------------------------------------------
    # R5: Mời cưới + BIDV là STK
    # ------------------------------------------------------------------
    cuoi_kw = ["đám cưới", "thiệp cưới", "lễ cưới", "hôn lễ",
               "mời cưới", "vu quy", "tân hôn", "wedding"]
    if any(k in all_text for k in cuoi_kw) and has_bidv(bidv_check_text):
        reasons.append("R5: Mời cưới + BIDV là STK")

    # ------------------------------------------------------------------
    # R6: Bán hàng + BIDV là STK (phải có dãy số 6+ chữ số)
    # ------------------------------------------------------------------
    banhang_kw = [
        "inbox", "order", "ship", "giao hàng", "mua hàng", "đặt hàng",
        "ck:", "ck :", "stk:", "stk :", "số tài khoản", "tài khoản ngân hàng",
    ]
    if any(k in all_text for k in banhang_kw) and has_bidv(bidv_check_text):
        if re.search(r"bidv.*?\d{6,}|\d{6,}.*?bidv", all_text):
            reasons.append("R6: Bán hàng + BIDV là STK")

    # ------------------------------------------------------------------
    # R7: Hướng dẫn bật kiếm tiền
    # ------------------------------------------------------------------
    kiemtien_kw = [
        "bật kiếm tiền", "kiếm tiền trên facebook", "kiếm tiền nội dung",
        "kiếm tiền từ nội dung",
        "#kiemtienfb", "professional mode", "professional dashboard",
        "bật ktnd", "kiếm tiền fb",
    ]
    if any(k in all_text for k in kiemtien_kw):
        reasons.append("R7: Hướng dẫn bật kiếm tiền")

    # ------------------------------------------------------------------
    # R8: Từ thiện/thiện nguyện + BIDV là STK
    # ------------------------------------------------------------------
    thiennguyen_kw = ["từ thiện", "thiện nguyện", "nam mô", "quyên góp",
                      "ủng hộ", "donate", "hảo tâm"]
    if any(k in all_text for k in thiennguyen_kw) \
            and (re.search(r"bidv.*?\d{6,}|\d{6,}.*?bidv", all_text)
                 or has_bidv(bidv_check_text)):
        reasons.append("R8: Từ thiện/thiện nguyện + BIDV là STK")

    # ------------------------------------------------------------------
    # R9: Đăng ký tham gia + BIDV là STK
    # ------------------------------------------------------------------
    if any(k in all_text for k in ["đăng ký tham gia", "đk tham gia",
                                    "form đăng ký", "đăng kí tham gia"]) \
            and has_bidv(bidv_check_text):
        reasons.append("R9: Đăng ký tham gia + BIDV là STK")

    # ------------------------------------------------------------------
    # R10: KPI mở tài khoản
    # ------------------------------------------------------------------
    r10_context = ["mở tài khoản", "mở tk", "inbox", "liên hệ", "nhắn tin", "zalo"]
    r10_direct  = any(k in all_text for k in ["hỗ trợ kpi", "kpi mở tài khoản", "nhận kpi"])
    r10_chaykpi = "chạy kpi" in all_text and any(k in all_text for k in r10_context)
    if r10_direct or r10_chaykpi:
        reasons.append("R10: Nhận hỗ trợ KPI mở tài khoản")

    # ------------------------------------------------------------------
    # R11: Chạy chỉ tiêu app
    # ------------------------------------------------------------------
    r11_context     = ["app", "mở tk", "mở tài khoản", "inbox", "liên hệ", "zalo"]
    r11_direct      = any(k in all_text for k in ["chỉ tiêu app", "chỉ tiêu mở tk"])
    r11_chaychitieu = "chạy chỉ tiêu" in all_text and any(k in all_text for k in r11_context)
    if r11_direct or r11_chaychitieu:
        reasons.append("R11: Chạy chỉ tiêu các app")

    # ------------------------------------------------------------------
    # R12: Xin tiền/donate + BIDV là STK
    # ------------------------------------------------------------------
    xintien_kw = [
        "xin tiền", "xin donate", "mạnh thường quân", "mạnh thương quân",
        "ủng hộ em", "giúp em với", "cầu xin", "thương giúp", "giúp đỡ",
    ]
    if any(k in all_text for k in xintien_kw) and has_bidv(bidv_check_text):
        reasons.append("R12: Xin tiền/donate + BIDV là STK")

    # ------------------------------------------------------------------
    # R13: Thêm ảnh + trống title/content
    # ------------------------------------------------------------------
    photo_kw = ["đã thêm ảnh", "add photo", "added a new photo",
                "added photos", "added a photo"]
    if any(k in all_text for k in photo_kw) \
            and not title.strip() and not content.strip():
        reasons.append("R13: Thêm ảnh + trống title/content")

    # ------------------------------------------------------------------
    # R14: BIC không liên quan bảo hiểm/NH/BIDV
    # ------------------------------------------------------------------
    check14 = all_text if is_post else tc_text
    if "bic" in check14 and not any(
        k in check14 for k in ["bảo hiểm", "insurance", "ngân hàng", "bank", "bidv"]
    ):
        reasons.append("R14: Nhắc BIC nhưng không liên quan bảo hiểm/NH/BIDV")

    # ------------------------------------------------------------------
    # R15: BID không liên quan chứng khoán/NH/BIDV
    # ------------------------------------------------------------------
    check15 = all_text if is_post else tc_text
    stock_kw = ["chứng khoán", "thị trường", "cổ phiếu", "stock",
                "securities", "ngân hàng", "bank"]
    if "bid" in check15 and "bidv" not in check15 \
            and not any(k in check15 for k in stock_kw + BIDV_KEYWORDS):
        reasons.append("R15: Nhắc BID nhưng không liên quan chứng khoán/NH/BIDV")

    # ------------------------------------------------------------------
    # R16: Nhả vía / Thả QR + BIDV
    # ------------------------------------------------------------------
    nha_via_kw = [
        "nhả vía", "nhả via", "nhả día", "nhả dia",
        "thả qr", "tha qr", "scan qr", "quét qr",
        "xin vía", "xin via", "chia vía", "chia via",
    ]
    if any(k in all_text for k in nha_via_kw) and has_bidv(all_text):
        reasons.append("R16: Nhả vía / Thả QR + BIDV")

    return reasons


# ---------------------------------------------------------------------------
# 4. PIPELINE: Load → Cascade → Label
# ---------------------------------------------------------------------------

def load_data(filepath: str | Path) -> pd.DataFrame:
    """Đọc file xlsx, kiểm tra cột bắt buộc, trả về DataFrame."""
    df = pd.read_excel(filepath, dtype=str)
    df = df.fillna("")

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"File thiếu cột: {missing}")

    return df


def apply_labels(df: pd.DataFrame) -> pd.DataFrame:
    """
    Áp dụng 16 rules + cascade logic, thêm 2 cột Spam_Label & Spam_Reason.
    Trả về DataFrame đã có nhãn (không thay đổi file gốc).
    """
    df = df.copy()

    # Bước 2: Xác định Post / Comment
    df["_is_post"] = df["Id"].astype(str) == df["ParentId"].astype(str)

    # Bước 5a: Đánh rule cho Posts
    spam_reasons: list[list[str]] = []
    for _, row in df.iterrows():
        if row["_is_post"]:
            spam_reasons.append(check_spam(row.to_dict(), is_post=True))
        else:
            spam_reasons.append([])  # placeholder, sẽ xử lý ở 5c

    df["_spam_reasons"] = spam_reasons

    # Bước 5b: Map post_id → spam_reasons
    post_spam_map: dict[str, list[str]] = {
        str(row["Id"]): row["_spam_reasons"]
        for _, row in df[df["_is_post"]].iterrows()
    }

    # Bước 5c: Xử lý Comments
    for idx, row in df[~df["_is_post"]].iterrows():
        parent_id     = str(row["ParentId"])
        parent_reasons = post_spam_map.get(parent_id)

        if parent_reasons is None:
            # Parent không có trong data → đánh rule độc lập
            df.at[idx, "_spam_reasons"] = check_spam(row.to_dict(), is_post=False)
        elif parent_reasons:
            # Post cha là SPAM → cascade
            df.at[idx, "_spam_reasons"] = [
                f"CASCADE: Post cha là spam ({parent_reasons[0]})"
            ]
        else:
            # Post cha NOT SPAM → comment NOT SPAM
            df.at[idx, "_spam_reasons"] = []

    # Bước 5d: Gán nhãn
    df["Spam_Label"]  = df["_spam_reasons"].apply(lambda r: "SPAM" if r else "NOT SPAM")
    df["Spam_Reason"] = df["_spam_reasons"].apply(lambda r: " | ".join(r) if r else "")

    # Xoá cột nội bộ
    df.drop(columns=["_is_post", "_spam_reasons"], inplace=True)

    return df


# ---------------------------------------------------------------------------
# 5. EXPORT: xlsx 4 sheets + format
# ---------------------------------------------------------------------------

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(color="FFFFFF", bold=True)
_SPAM_FILL    = PatternFill("solid", fgColor="FFE0E0")
_SPAM_FONT    = Font(color="CC0000", bold=True)


def _style_sheet(ws, df_sheet: pd.DataFrame, mark_spam: bool = False) -> None:
    """Áp dụng header style, highlight SPAM, freeze pane, autofilter, auto-width."""
    # Header
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Highlight SPAM rows
    if mark_spam and "Spam_Label" in df_sheet.columns:
        spam_col_idx = list(df_sheet.columns).index("Spam_Label") + 1
        for row_idx, val in enumerate(df_sheet["Spam_Label"], start=2):
            if val == "SPAM":
                for col_idx in range(1, len(df_sheet.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = _SPAM_FILL
                    if col_idx == spam_col_idx:
                        cell.font = _SPAM_FONT

    # Freeze pane tại A2
    ws.freeze_panes = "A2"

    # AutoFilter
    ws.auto_filter.ref = ws.dimensions

    # Auto column width (max 60)
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(_safe_str(c.value)) for c in col_cells), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def export_result(df: pd.DataFrame, output_path: str | Path) -> None:
    """
    Xuất DataFrame đã nhãn ra file xlsx 4 sheets:
      - All Data  : toàn bộ dữ liệu, highlight SPAM
      - SPAM      : chỉ dòng SPAM
      - NOT SPAM  : chỉ dòng NOT SPAM
      - Summary   : tóm tắt thống kê
    """
    df_spam     = df[df["Spam_Label"] == "SPAM"].copy()
    df_not_spam = df[df["Spam_Label"] == "NOT SPAM"].copy()

    # ---- Summary ----
    spam_count     = len(df_spam)
    not_spam_count = len(df_not_spam)
    total          = len(df)
    spam_pct       = spam_count / total * 100 if total else 0

    # Thống kê theo rule
    all_reasons: list[str] = []
    for reason_str in df_spam["Spam_Reason"]:
        all_reasons.extend(r.split(":")[0].strip() for r in reason_str.split(" | ") if r)

    from collections import Counter
    rule_counts = Counter(all_reasons)
    summary_rows = [
        {"Chỉ số": "Tổng dòng",     "Giá trị": total},
        {"Chỉ số": "SPAM",          "Giá trị": spam_count},
        {"Chỉ số": "NOT SPAM",      "Giá trị": not_spam_count},
        {"Chỉ số": "% SPAM",        "Giá trị": f"{spam_pct:.1f}%"},
        {"Chỉ số": "---",           "Giá trị": "--- Chi tiết theo Rule ---"},
        *[{"Chỉ số": rule, "Giá trị": cnt}
          for rule, cnt in sorted(rule_counts.items())],
    ]
    df_summary = pd.DataFrame(summary_rows)

    # Cảnh báo nếu > 40%
    if spam_pct > 40:
        print(f"⚠️  CẢNH BÁO: Tỉ lệ SPAM = {spam_pct:.1f}% (> 40%) — kiểm tra lại rules!")

    # Ghi file
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer,         sheet_name="All Data",  index=False)
        df_spam.to_excel(writer,    sheet_name="SPAM",      index=False)
        df_not_spam.to_excel(writer,sheet_name="NOT SPAM",  index=False)
        df_summary.to_excel(writer, sheet_name="Summary",   index=False)

    # Styling (openpyxl sau khi đóng ExcelWriter)
    wb = load_workbook(str(output_path))
    _style_sheet(wb["All Data"],  df,          mark_spam=True)
    _style_sheet(wb["SPAM"],      df_spam,     mark_spam=True)
    _style_sheet(wb["NOT SPAM"],  df_not_spam, mark_spam=False)
    _style_sheet(wb["Summary"],   df_summary,  mark_spam=False)
    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# 6. PUBLIC API
# ---------------------------------------------------------------------------

def run(input_path: str | Path, output_path: str | Path | None = None) -> pd.DataFrame:
    """
    Hàm chạy toàn bộ pipeline.

    Parameters
    ----------
    input_path  : đường dẫn file xlsx đầu vào
    output_path : đường dẫn file xlsx đầu ra (mặc định: <input>_spam_labeled.xlsx)

    Returns
    -------
    DataFrame đã có cột Spam_Label & Spam_Reason
    """
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.parent / f"{input_path.stem}_spam_labeled.xlsx"
    output_path = Path(output_path)

    print(f"📂 Đọc file: {input_path}")
    df = load_data(input_path)
    print(f"   → {len(df):,} dòng | {df.columns.tolist()}")

    print("⚙️  Áp dụng 16 rules + cascade logic...")
    df_labeled = apply_labels(df)

    spam_count = (df_labeled["Spam_Label"] == "SPAM").sum()
    total      = len(df_labeled)
    print(f"   → SPAM: {spam_count:,} / {total:,} ({spam_count/total*100:.1f}%)")

    print(f"💾 Xuất file: {output_path}")
    export_result(df_labeled, output_path)
    print("✅ Hoàn tất!")

    return df_labeled


# ---------------------------------------------------------------------------
# 7. CLI ENTRY POINT
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Lọc spam Social Listening BIDV (v1.4)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("input",  help="File xlsx đầu vào")
    p.add_argument("--output", "-o", default=None,
                   help="File xlsx đầu ra (mặc định: <input>_spam_labeled.xlsx)")
    return p


def main(argv: list[str] | None = None) -> None:
    args = _build_parser().parse_args(argv)
    try:
        run(args.input, args.output)
    except FileNotFoundError as e:
        print(f"❌ Không tìm thấy file: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"❌ Lỗi dữ liệu: {e}", file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()